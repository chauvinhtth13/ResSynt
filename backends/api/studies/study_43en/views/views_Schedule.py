import logging
from datetime import date, datetime, timedelta

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q, Case, When, Value, IntegerField
from django.utils.translation import gettext as _
from django.utils.timezone import localtime
from backends.api.studies.study_43en.services.context_processors import upcoming_appointments
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Import models từ study app
from backends.studies.study_43en.models.patient import (
    FU_CASE_28, FU_CASE_90, SAM_CASE, ENR_CASE
)
from backends.studies.study_43en.models.contact import (
    FU_CONTACT_28, FU_CONTACT_90, ENR_CONTACT
)
from backends.studies.study_43en.models.schedule import (
    FollowUpStatus
)

#  Import site utilities
from backends.studies.study_43en.utils.site_utils import (
    get_site_filter_params,
    get_filtered_queryset
)

logger = logging.getLogger(__name__)


@login_required
def followup_tracking_list(request):
    """Hiển thị danh sách theo dõi lịch hẹn - 3 TABLES: pending/completed/missed"""
    
    study_db = getattr(request, 'study_db_alias', 'db_study_43en')
    today = date.today()
    
    day_of_week = today.weekday()
    start_of_week = today - timedelta(days=day_of_week)
    end_of_week = start_of_week + timedelta(days=6)
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from and not date_to:
        date_from = start_of_week
        date_to = end_of_week
    else:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        except ValueError:
            date_from = start_of_week
            date_to = end_of_week
    
    visit_filter = request.GET.get('visit')
    status_filter = request.GET.get('status')
    subject_type_filter = request.GET.get('subject_type')
    search_query = request.GET.get('search')
    
    site_filter, filter_type = get_site_filter_params(request)
    logger.info(f"followup_tracking_list - User: {request.user.username}, Site: {site_filter}, Type: {filter_type}")
    
    # Auto-update outdated statuses
    try:
        pending_to_update = get_filtered_queryset(FollowUpStatus, site_filter, filter_type).exclude(
            STATUS__in=['COMPLETED', 'MISSED']
        )
        updated = 0
        for followup in pending_to_update:
            old_status = followup.STATUS
            
            if followup.ACTUAL_DATE:
                followup.STATUS = 'COMPLETED'
            elif followup.MISSED_DATE:
                followup.STATUS = 'MISSED'
            elif followup.EXPECTED_TO and today > followup.EXPECTED_TO:
                followup.STATUS = 'LATE'
            elif followup.EXPECTED_DATE and today > followup.EXPECTED_DATE:
                followup.STATUS = 'LATE'
            
            if followup.STATUS != old_status:
                followup.save(using=study_db, update_fields=['STATUS'])
                updated += 1
        
        if updated > 0:
            logger.info(f"Auto-updated {updated} followup statuses")
    except Exception as e:
        logger.error(f"Error auto-updating status: {e}")
    
    base_followups = get_filtered_queryset(FollowUpStatus, site_filter, filter_type)
    
    if date_from:
        base_followups = base_followups.filter(
            Q(EXPECTED_DATE__gte=date_from) | 
            Q(EXPECTED_FROM__gte=date_from) |
            Q(ACTUAL_DATE__gte=date_from) |
            Q(MISSED_DATE__gte=date_from)
        )
    if date_to:
        base_followups = base_followups.filter(
            Q(EXPECTED_DATE__lte=date_to) | 
            Q(EXPECTED_TO__lte=date_to) |
            Q(ACTUAL_DATE__lte=date_to) |
            Q(MISSED_DATE__lte=date_to)
        )
    
    if visit_filter:
        base_followups = base_followups.filter(VISIT=visit_filter)
    if subject_type_filter:
        base_followups = base_followups.filter(SUBJECT_TYPE=subject_type_filter)
    if search_query:
        base_followups = base_followups.filter(
            Q(USUBJID__icontains=search_query) | 
            Q(INITIAL__icontains=search_query) |
            Q(PHONE__icontains=search_query)
        )
    
    pending_followups = base_followups.exclude(STATUS__in=['COMPLETED', 'MISSED']).order_by(
        Case(
            When(STATUS='LATE', then=Value(0)),
            When(STATUS='UPCOMING', then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        ),
        'EXPECTED_DATE'
    )
    
    completed_followups = base_followups.filter(STATUS='COMPLETED').order_by('-ACTUAL_DATE')
    missed_followups = base_followups.filter(STATUS='MISSED').order_by('-MISSED_DATE')
    
    if status_filter:
        if status_filter == 'COMPLETED':
            pending_followups = pending_followups.none()
            missed_followups = missed_followups.none()
        elif status_filter == 'MISSED':
            pending_followups = pending_followups.none()
            completed_followups = completed_followups.none()
        else:
            completed_followups = completed_followups.none()
            missed_followups = missed_followups.none()
    
    total_appointments = base_followups.count()
    completed_count = completed_followups.count()
    late_count = pending_followups.filter(STATUS='LATE').count()
    missed_count = missed_followups.count()
    upcoming_count = pending_followups.filter(STATUS='UPCOMING').count()
    
    pending_paginator = Paginator(pending_followups, 15)
    completed_paginator = Paginator(completed_followups, 10)
    missed_paginator = Paginator(missed_followups, 10)
    
    pending_page = request.GET.get('pending_page', 1)
    completed_page = request.GET.get('completed_page', 1)
    missed_page = request.GET.get('missed_page', 1)
    
    try:
        pending_page_obj = pending_paginator.page(pending_page)
    except (PageNotAnInteger, EmptyPage):
        pending_page_obj = pending_paginator.page(1)
    
    try:
        completed_page_obj = completed_paginator.page(completed_page)
    except (PageNotAnInteger, EmptyPage):
        completed_page_obj = completed_paginator.page(1)
    
    try:
        missed_page_obj = missed_paginator.page(missed_page)
    except (PageNotAnInteger, EmptyPage):
        missed_page_obj = missed_paginator.page(1)
    
    upcoming_stats = {
        'PATIENT': {'V2': 0, 'V3': 0, 'V4': 0, 'total': 0},
        'CONTACT': {'V2': 0, 'V3': 0, 'V4': 0, 'total': 0}
    }
    
    upcoming_date = today + timedelta(days=7)
    
    upcoming_followups = get_filtered_queryset(FollowUpStatus, site_filter, filter_type).filter(
        STATUS='UPCOMING',
        EXPECTED_DATE__lte=upcoming_date,
        EXPECTED_DATE__gte=today
    )
    
    for followup in upcoming_followups:
        upcoming_stats[followup.SUBJECT_TYPE][followup.VISIT] = upcoming_stats[followup.SUBJECT_TYPE].get(followup.VISIT, 0) + 1
        upcoming_stats[followup.SUBJECT_TYPE]['total'] = upcoming_stats[followup.SUBJECT_TYPE].get('total', 0) + 1
    
    late_stats = {
        'PATIENT': {'V2': 0, 'V3': 0, 'V4': 0, 'total': 0},
        'CONTACT': {'V2': 0, 'V3': 0, 'V4': 0, 'total': 0}
    }
    
    late_followups = get_filtered_queryset(FollowUpStatus, site_filter, filter_type).filter(STATUS='LATE')
    
    for followup in late_followups:
        late_stats[followup.SUBJECT_TYPE][followup.VISIT] = late_stats[followup.SUBJECT_TYPE].get(followup.VISIT, 0) + 1
        late_stats[followup.SUBJECT_TYPE]['total'] = late_stats[followup.SUBJECT_TYPE].get('total', 0) + 1
    
    context = {
        'pending_followups': pending_page_obj,
        'completed_followups': completed_page_obj,
        'missed_followups': missed_page_obj,
        'visit_choices': FollowUpStatus.VISIT_CHOICES,
        'status_choices': FollowUpStatus.STATUS_CHOICES,
        'subject_type_choices': FollowUpStatus.SUBJECT_TYPE_CHOICES,
        'current_visit': visit_filter,
        'current_status': status_filter,
        'current_subject_type': subject_type_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'upcoming_stats': upcoming_stats,
        'late_stats': late_stats,
        'total_appointments': total_appointments,
        'completed_count': completed_count,
        'late_count': late_count,
        'missed_count': missed_count,
        'upcoming_count': upcoming_count,
    }
    
    return render(request, 'studies/study_43en/base/followup_tracking_list.html', context)


@login_required
def complete_followup_ajax(request, pk):
    """
     FIXED: Mark follow-up as completed and sync to CRF
    Two-way sync: FollowUpStatus ↔ CRF Forms
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST allowed'}, status=405)
    
    study_db = getattr(request, 'study_db_alias', 'db_study_43en')
    
    try:
        # Get follow-up record
        followup = FollowUpStatus.objects.using(study_db).get(pk=pk)
        
        # Get actual date from request
        actual_date_str = request.POST.get('actual_date')
        if not actual_date_str:
            return JsonResponse({
                'success': False, 
                'message': 'Ngày hoàn thành không được để trống'
            }, status=400)
        
        try:
            actual_date = datetime.strptime(actual_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Định dạng ngày không hợp lệ. Sử dụng YYYY-MM-DD'
            }, status=400)
        
        #  UPDATE FollowUpStatus FIRST
        followup.ACTUAL_DATE = actual_date
        followup.STATUS = 'COMPLETED'
        followup.save(using=study_db)
        
        logger.info(f" Marked followup {pk} ({followup.USUBJID} {followup.VISIT}) as COMPLETED on {actual_date}")
        
        #  SYNC TO CRF FORM - FIXED LOOKUP
        sync_success = False
        crf_form_name = None
        sync_message = ""
        
        try:
            if followup.SUBJECT_TYPE == 'PATIENT':
                #  Get ENR_CASE by USUBJID string
                try:
                    enr_case = ENR_CASE.objects.using(study_db).select_related('USUBJID').get(
                        USUBJID__USUBJID=followup.USUBJID  # SCR_CASE.USUBJID = followup.USUBJID
                    )
                    logger.info(f" Found ENR_CASE for {followup.USUBJID}: {enr_case.pk}")
                except ENR_CASE.DoesNotExist:
                    logger.error(f" ENR_CASE not found for {followup.USUBJID}")
                    return JsonResponse({
                        'success': True,  # Still mark as completed
                        'message': f'Đã hoàn thành nhưng chưa có bệnh nhân enrollment cho {followup.USUBJID}',
                        'data': {
                            'followup_id': followup.pk,
                            'usubjid': followup.USUBJID,
                            'visit': followup.get_VISIT_display(),
                            'actual_date': actual_date.strftime('%d/%m/%Y'),
                            'status': 'COMPLETED',
                            'sync_success': False,
                            'sync_message': 'Chưa có enrollment'
                        }
                    })
                
                if followup.VISIT == 'V2':
                    # V2 = Sample collection
                    try:
                        sample = SAM_CASE.objects.using(study_db).get(
                            USUBJID=enr_case,
                            SAMPLE_TYPE='2'
                        )
                        sample.SAMPLE = True
                        # Update collection dates
                        if sample.STOOL and not sample.STOOLDATE:
                            sample.STOOLDATE = actual_date
                        if sample.RECTSWAB and not sample.RECTSWABDATE:
                            sample.RECTSWABDATE = actual_date
                        if sample.THROATSWAB and not sample.THROATSWABDATE:
                            sample.THROATSWABDATE = actual_date
                        if sample.BLOOD and not sample.BLOODDATE:
                            sample.BLOODDATE = actual_date
                        sample.save(using=study_db)
                        sync_success = True
                        crf_form_name = 'SAM_CASE (V2)'
                        sync_message = f"Đã cập nhật mẫu V2"
                        logger.info(f" Synced to SAM_CASE for {followup.USUBJID}")
                    except SAM_CASE.DoesNotExist:
                        sync_message = "Chưa có form SAM_CASE V2"
                        logger.warning(f" SAM_CASE V2 not found for {followup.USUBJID}")
                
                elif followup.VISIT == 'V3':
                    # V3 = FU_CASE_28
                    try:
                        fu_case = FU_CASE_28.objects.using(study_db).get(USUBJID=enr_case)
                        fu_case.EvaluateDate = actual_date
                        fu_case.EvaluatedAtDay28 = 'Yes'
                        fu_case.save(using=study_db, update_fields=['EvaluateDate', 'EvaluatedAtDay28'])
                        sync_success = True
                        crf_form_name = 'FU_CASE_28'
                        sync_message = f"Đã cập nhật FU_CASE_28"
                        logger.info(f" Synced to FU_CASE_28 for {followup.USUBJID}: EvaluateDate={actual_date}")
                    except FU_CASE_28.DoesNotExist:
                        #  TRY TO CREATE FU_CASE_28
                        try:
                            fu_case = FU_CASE_28.objects.using(study_db).create(
                                USUBJID=enr_case,
                                EvaluateDate=actual_date,
                                EvaluatedAtDay28='Yes'
                            )
                            sync_success = True
                            crf_form_name = 'FU_CASE_28 (Created)'
                            sync_message = f"Đã tạo mới FU_CASE_28"
                            logger.info(f" Created FU_CASE_28 for {followup.USUBJID}: EvaluateDate={actual_date}")
                        except Exception as create_error:
                            sync_message = f"Không thể tạo FU_CASE_28: {str(create_error)}"
                            logger.error(f" Error creating FU_CASE_28: {create_error}", exc_info=True)
                    except Exception as e:
                        sync_message = f"Lỗi sync FU_CASE_28: {str(e)}"
                        logger.error(f" Error syncing FU_CASE_28: {e}", exc_info=True)
                
                elif followup.VISIT == 'V4':
                    # V4 = FU_CASE_90
                    try:
                        fu_case = FU_CASE_90.objects.using(study_db).get(USUBJID=enr_case)
                        fu_case.EvaluateDate = actual_date
                        fu_case.EvaluatedAtDay90 = 'Yes'
                        fu_case.save(using=study_db, update_fields=['EvaluateDate', 'EvaluatedAtDay90'])
                        sync_success = True
                        crf_form_name = 'FU_CASE_90'
                        sync_message = f"Đã cập nhật FU_CASE_90"
                        logger.info(f" Synced to FU_CASE_90 for {followup.USUBJID}")
                    except FU_CASE_90.DoesNotExist:
                        #  TRY TO CREATE FU_CASE_90
                        try:
                            fu_case = FU_CASE_90.objects.using(study_db).create(
                                USUBJID=enr_case,
                                EvaluateDate=actual_date,
                                EvaluatedAtDay90='Yes'
                            )
                            sync_success = True
                            crf_form_name = 'FU_CASE_90 (Created)'
                            sync_message = f"Đã tạo mới FU_CASE_90"
                            logger.info(f" Created FU_CASE_90 for {followup.USUBJID}")
                        except Exception as create_error:
                            sync_message = f"Không thể tạo FU_CASE_90: {str(create_error)}"
                            logger.error(f" Error creating FU_CASE_90: {create_error}", exc_info=True)
            
            elif followup.SUBJECT_TYPE == 'CONTACT':
                #  Get ENR_CONTACT by USUBJID string
                try:
                    enr_contact = ENR_CONTACT.objects.using(study_db).select_related('USUBJID').get(
                        USUBJID__USUBJID=followup.USUBJID
                    )
                    logger.info(f" Found ENR_CONTACT for {followup.USUBJID}")
                except ENR_CONTACT.DoesNotExist:
                    logger.error(f" ENR_CONTACT not found for {followup.USUBJID}")
                    return JsonResponse({
                        'success': True,
                        'message': f'Đã hoàn thành nhưng chưa có contact enrollment cho {followup.USUBJID}',
                        'data': {
                            'followup_id': followup.pk,
                            'usubjid': followup.USUBJID,
                            'visit': followup.get_VISIT_display(),
                            'actual_date': actual_date.strftime('%d/%m/%Y'),
                            'status': 'COMPLETED',
                            'sync_success': False,
                            'sync_message': 'Chưa có enrollment'
                        }
                    })
                
                if followup.VISIT == 'V2':
                    # Contact V2 = FU_CONTACT_28
                    try:
                        fu_contact = FU_CONTACT_28.objects.using(study_db).get(USUBJID=enr_contact)
                        fu_contact.EvaluateDate = actual_date
                        fu_contact.EvaluatedAtDay28 = 'Yes'
                        fu_contact.save(using=study_db, update_fields=['EvaluateDate', 'EvaluatedAtDay28'])
                        sync_success = True
                        crf_form_name = 'FU_CONTACT_28'
                        sync_message = f"Đã cập nhật FU_CONTACT_28"
                        logger.info(f" Synced to FU_CONTACT_28 for {followup.USUBJID}")
                    except FU_CONTACT_28.DoesNotExist:
                        try:
                            fu_contact = FU_CONTACT_28.objects.using(study_db).create(
                                USUBJID=enr_contact,
                                EvaluateDate=actual_date,
                                EvaluatedAtDay28='Yes'
                            )
                            sync_success = True
                            crf_form_name = 'FU_CONTACT_28 (Created)'
                            sync_message = f"Đã tạo mới FU_CONTACT_28"
                            logger.info(f" Created FU_CONTACT_28 for {followup.USUBJID}")
                        except Exception as create_error:
                            sync_message = f"Không thể tạo FU_CONTACT_28: {str(create_error)}"
                            logger.error(f" Error creating FU_CONTACT_28: {create_error}", exc_info=True)
                
                elif followup.VISIT == 'V3':
                    # Contact V3 = FU_CONTACT_90
                    try:
                        fu_contact = FU_CONTACT_90.objects.using(study_db).get(USUBJID=enr_contact)
                        fu_contact.EvaluateDate = actual_date
                        fu_contact.EvaluatedAtDay90 = 'Yes'
                        fu_contact.save(using=study_db, update_fields=['EvaluateDate', 'EvaluatedAtDay90'])
                        sync_success = True
                        crf_form_name = 'FU_CONTACT_90'
                        sync_message = f"Đã cập nhật FU_CONTACT_90"
                        logger.info(f" Synced to FU_CONTACT_90 for {followup.USUBJID}")
                    except FU_CONTACT_90.DoesNotExist:
                        try:
                            fu_contact = FU_CONTACT_90.objects.using(study_db).create(
                                USUBJID=enr_contact,
                                EvaluateDate=actual_date,
                                EvaluatedAtDay90='Yes'
                            )
                            sync_success = True
                            crf_form_name = 'FU_CONTACT_90 (Created)'
                            sync_message = f"Đã tạo mới FU_CONTACT_90"
                            logger.info(f" Created FU_CONTACT_90 for {followup.USUBJID}")
                        except Exception as create_error:
                            sync_message = f"Không thể tạo FU_CONTACT_90: {str(create_error)}"
                            logger.error(f" Error creating FU_CONTACT_90: {create_error}", exc_info=True)
        
        except Exception as e:
            sync_message = f"Lỗi khi đồng bộ: {str(e)}"
            logger.error(f" Unexpected error in sync: {e}", exc_info=True)
        
        # Build response
        response_data = {
            'success': True,
            'message': 'Đã đánh dấu hoàn thành thành công!',
            'data': {
                'followup_id': followup.pk,
                'usubjid': followup.USUBJID,
                'visit': followup.get_VISIT_display(),
                'actual_date': actual_date.strftime('%d/%m/%Y'),
                'status': 'COMPLETED',
                'sync_success': sync_success,
                'crf_form': crf_form_name,
                'sync_message': sync_message
            }
        }
        
        if sync_success:
            response_data['message'] += f' {sync_message}.'
        else:
            response_data['message'] += f' {sync_message}.'
        
        return JsonResponse(response_data)
        
    except FollowUpStatus.DoesNotExist:
        logger.error(f" FollowUpStatus {pk} not found")
        return JsonResponse({
            'success': False,
            'message': 'Không tìm thấy lịch hẹn'
        }, status=404)
    except Exception as e:
        logger.error(f" Error completing followup {pk}: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@login_required
def update_followup_status(request, pk):
    """Cập nhật trạng thái theo dõi"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Chỉ hỗ trợ phương thức POST'})
    
    #  Get study database
    study_db = getattr(request, 'study_db_alias', 'db_study_43en')
    
    try:
        #  Query từ study database
        followup = FollowUpStatus.objects.using(study_db).get(pk=pk)
        
        # Cập nhật từ form
        actual_date = request.POST.get('actual_date')
        status = request.POST.get('status')
        notes = request.POST.get('notes')
        
        if actual_date:
            followup.ACTUAL_DATE = datetime.strptime(actual_date, '%Y-%m-%d').date()
            
            # Cập nhật form tương ứng nếu có
            if followup.SUBJECT_TYPE == 'PATIENT':
                if followup.VISIT == 'V3':
                    try:
                        fu_case = FU_CASE_28.objects.using(study_db).get(USUBJID_id=followup.USUBJID)
                        fu_case.ASSESSDATE = followup.ACTUAL_DATE
                        fu_case.save(using=study_db, update_fields=['ASSESSDATE'])
                    except FU_CASE_28.DoesNotExist:
                        pass
                        
                elif followup.VISIT == 'V4':
                    try:
                        fu_case = FU_CASE_90.objects.using(study_db).get(USUBJID_id=followup.USUBJID)
                        fu_case.ASSESSDATE = followup.ACTUAL_DATE
                        fu_case.save(using=study_db, update_fields=['ASSESSDATE'])
                    except FU_CASE_90.DoesNotExist:
                        pass
            
            elif followup.SUBJECT_TYPE == 'CONTACT':
                if followup.VISIT == 'V2':
                    try:
                        fu_case = FU_CONTACT_28.objects.using(study_db).get(USUBJID_id=followup.USUBJID)
                        fu_case.ASSESSDATE = followup.ACTUAL_DATE
                        fu_case.save(using=study_db, update_fields=['ASSESSDATE'])
                    except FU_CONTACT_28.DoesNotExist:
                        pass
                        
                elif followup.VISIT == 'V3':
                    try:
                        fu_case = FU_CONTACT_90.objects.using(study_db).get(USUBJID_id=followup.USUBJID)
                        fu_case.ASSESSDATE = followup.ACTUAL_DATE
                        fu_case.save(using=study_db, update_fields=['ASSESSDATE'])
                    except FU_CONTACT_90.DoesNotExist:
                        pass
            
        if status:
            followup.STATUS = status
            
        if notes:
            followup.NOTES = notes
            
        followup.save(using=study_db)
        
        return JsonResponse({
            'success': True,
            'message': 'Đã cập nhật thành công',
            'status': followup.get_STATUS_display(),
            'actual_date': followup.ACTUAL_DATE.strftime('%d/%m/%Y') if followup.ACTUAL_DATE else None
        })
        
    except FollowUpStatus.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy bản ghi'}, status=404)
    except Exception as e:
        logger.error(f"Error updating followup status: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def export_followup_tracking(request):
    """Xuất danh sách theo dõi lịch hẹn ra Excel -  WITH SITE FILTERING"""
    
    #  Get study database
    study_db = getattr(request, 'study_db_alias', 'db_study_43en')
    
    #  NEW: Get site filter with proper strategy
    site_filter, filter_type = get_site_filter_params(request)
    
    logger.info(f"export_followup_tracking - User: {request.user.username}, Site: {site_filter}, Type: {filter_type}")
    
    # Lấy các tham số
    visit_filter = request.GET.get('visit')
    status_filter = request.GET.get('status')
    subject_type_filter = request.GET.get('subject_type')
    search_query = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    #  Query với site filtering
    followups = get_filtered_queryset(FollowUpStatus, site_filter, filter_type)
    
    # Áp dụng filter theo khoảng ngày
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            followups = followups.filter(Q(EXPECTED_DATE__gte=date_from) | 
                                       Q(EXPECTED_FROM__gte=date_from) |
                                       Q(ACTUAL_DATE__gte=date_from))
        except ValueError:
            pass
            
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            followups = followups.filter(Q(EXPECTED_DATE__lte=date_to) | 
                                       Q(EXPECTED_TO__lte=date_to) |
                                       Q(ACTUAL_DATE__lte=date_to))
        except ValueError:
            pass
    
    # Áp dụng các bộ lọc
    if visit_filter:
        followups = followups.filter(VISIT=visit_filter)
    
    if status_filter:
        followups = followups.filter(STATUS=status_filter)
    
    if subject_type_filter:
        followups = followups.filter(SUBJECT_TYPE=subject_type_filter)
    
    if search_query:
        followups = followups.filter(
            Q(USUBJID__icontains=search_query) | 
            Q(INITIAL__icontains=search_query)
        )
    
    # Sắp xếp
    followups = followups.order_by(
        Case(
            When(STATUS='LATE', then=Value(0)),
            When(STATUS='UPCOMING', then=Value(1)),
            When(STATUS='MISSED', then=Value(2)),
            When(STATUS='COMPLETED', then=Value(3)),
            default=Value(4),
            output_field=IntegerField(),
        ),
        'EXPECTED_DATE'
    )
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách theo dõi"
    
    # Thêm header
    headers = [
        'STT', 'USUBJID', 'Tên viết tắt', 'Loại đối tượng', 'Lần thăm', 
        'Ngày dự kiến', 'Khoảng thời gian', 'Ngày thực tế', 'Trạng thái', 'SĐT'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Thêm dữ liệu
    status_map = dict(FollowUpStatus.STATUS_CHOICES)
    subject_type_map = dict(FollowUpStatus.SUBJECT_TYPE_CHOICES)
    visit_map = dict(FollowUpStatus.VISIT_CHOICES)
    
    for row_num, followup in enumerate(followups, 2):
        ws.cell(row=row_num, column=1, value=row_num-1)
        ws.cell(row=row_num, column=2, value=followup.USUBJID)
        ws.cell(row=row_num, column=3, value=followup.INITIAL)
        ws.cell(row=row_num, column=4, value=subject_type_map.get(followup.SUBJECT_TYPE, followup.SUBJECT_TYPE))
        ws.cell(row=row_num, column=5, value=visit_map.get(followup.VISIT, followup.VISIT))
        
        if followup.EXPECTED_DATE:
            ws.cell(row=row_num, column=6, value=followup.EXPECTED_DATE)
        
        expected_range = ""
        if followup.EXPECTED_FROM and followup.EXPECTED_TO:
            expected_range = f"{followup.EXPECTED_FROM.strftime('%d/%m/%Y')} - {followup.EXPECTED_TO.strftime('%d/%m/%Y')}"
        ws.cell(row=row_num, column=7, value=expected_range)
        
        if followup.ACTUAL_DATE:
            ws.cell(row=row_num, column=8, value=followup.ACTUAL_DATE)
        
        ws.cell(row=row_num, column=9, value=status_map.get(followup.STATUS, followup.STATUS))
        ws.cell(row=row_num, column=10, value=followup.PHONE)
    
    # Tự động điều chỉnh chiều rộng cột
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Tạo response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=theo_doi_lich_hen_{localtime().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response



# ==========================================
# NOTIFICATION API ENDPOINTS
# ==========================================

@login_required
def mark_notification_read(request):
    """
     Đánh dấu 1 notification đã đọc
    POST: /studies/43en/api/notification/read/
    Body: {"notif_id": "003-A-001_V2_20251110"}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        notif_id = data.get('notif_id')
        
        if not notif_id:
            return JsonResponse({'success': False, 'message': 'Missing notif_id'}, status=400)
        
        #  Lấy danh sách đã đọc từ session
        read_notifications = request.session.get('read_notifications', [])
        
        #  Thêm vào nếu chưa có
        if notif_id not in read_notifications:
            read_notifications.append(notif_id)
            request.session['read_notifications'] = read_notifications
            request.session.modified = True
        
        #  Đếm lại số thông báo chưa đọc
        context = upcoming_appointments(request)
        
        return JsonResponse({
            'success': True,
            'message': 'Đã đánh dấu đọc',
            'unread_count': context['unread_count']
        })
        
    except Exception as e:
        logger.error(f"Error marking notification read: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def mark_all_notifications_read(request):
    """
     Đánh dấu TẤT CẢ notifications đã đọc - WITH SITE FILTERING
    POST: /studies/43en/api/notification/read-all/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST allowed'}, status=405)
    
    try:
        from datetime import date, timedelta
        from backends.studies.study_43en.models.schedule import FollowUpStatus
        
        today = date.today()
        upcoming_date = today + timedelta(days=3)
        
        #  NEW: Get site filter với 3 strategies
        site_filter, filter_type = get_site_filter_params(request)
        
        #  Lấy TẤT CẢ upcoming followups với site filtering
        upcoming_followups = get_filtered_queryset(FollowUpStatus, site_filter, filter_type).filter(
            EXPECTED_DATE__gte=today,
            EXPECTED_DATE__lte=upcoming_date,
            STATUS__in=['UPCOMING', 'LATE']
        )
        
        #  Tạo danh sách notif_id
        read_notifications = []
        for followup in upcoming_followups:
            notif_id = f"{followup.USUBJID}_{followup.VISIT}_{followup.EXPECTED_DATE.strftime('%Y%m%d')}"
            read_notifications.append(notif_id)
        
        #  Lưu vào session
        request.session['read_notifications'] = read_notifications
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'message': f'Đã đánh dấu {len(read_notifications)} thông báo đã đọc',
            'unread_count': 0,
            'marked_count': len(read_notifications)
        })
        
    except Exception as e:
        logger.error(f"Error marking all notifications read: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def get_notification_count(request):
    """
     API để lấy số thông báo chưa đọc (dùng cho polling/refresh)
    GET: /studies/43en/api/notification/count/
    """
    try:
        context = upcoming_appointments(request)
        
        return JsonResponse({
            'success': True,
            'total_count': context['upcoming_count'],
            'unread_count': context['unread_count']
        })
        
    except Exception as e:
        logger.error(f"Error getting notification count: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    

@login_required
def mark_followup_missed(request, pk):
    """
     Mark follow-up as MISSED
    POST: /studies/43en/followup-tracking/<pk>/missed/
    Body: {"missed_date": "2024-07-15", "missed_reason": "Không liên lạc được"}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST allowed'}, status=405)
    
    study_db = getattr(request, 'study_db_alias', 'db_study_43en')
    
    try:
        followup = FollowUpStatus.objects.using(study_db).get(pk=pk)
        
        # Get missed date
        missed_date_str = request.POST.get('missed_date')
        if not missed_date_str:
            return JsonResponse({
                'success': False,
                'message': 'Ngày missed không được để trống'
            }, status=400)
        
        try:
            missed_date = datetime.strptime(missed_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Định dạng ngày không hợp lệ. Sử dụng YYYY-MM-DD'
            }, status=400)
        
        # Get reason
        missed_reason = request.POST.get('missed_reason', '').strip()
        if not missed_reason:
            return JsonResponse({
                'success': False,
                'message': 'Lý do missed không được để trống'
            }, status=400)
        
        #  UPDATE Status
        followup.MISSED_DATE = missed_date
        followup.MISSED_REASON = missed_reason
        followup.STATUS = 'MISSED'
        followup.save(using=study_db)
        
        logger.info(f" Marked followup {pk} ({followup.USUBJID} {followup.VISIT}) as MISSED on {missed_date}")
        
        return JsonResponse({
            'success': True,
            'message': 'Đã đánh dấu Missed thành công!',
            'data': {
                'followup_id': followup.pk,
                'usubjid': followup.USUBJID,
                'visit': followup.get_VISIT_display(),
                'missed_date': missed_date.strftime('%d/%m/%Y'),
                'missed_reason': missed_reason,
                'status': 'MISSED'
            }
        })
        
    except FollowUpStatus.DoesNotExist:
        logger.error(f" FollowUpStatus {pk} not found")
        return JsonResponse({
            'success': False,
            'message': 'Không tìm thấy lịch hẹn'
        }, status=404)
    except Exception as e:
        logger.error(f" Error marking followup {pk} as missed: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Lỗi: {str(e)}'
        }, status=500)
